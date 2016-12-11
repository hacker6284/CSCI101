from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook('list.xlsx')
ws = wb.active
inp = []
contenders = []
#while "done" not in inp:
#	 inp = [raw_input("Enter the Contender: ")]
#	 if "done" not in inp:
#		 contenders += [inp]
col = ws['A']
for cell in col:
	inp = [str(cell.value)]
	contenders += [inp]
wb = Workbook()
ws = wb.active
length = len(contenders)
j = 0
changed = 1
while changed == 1:
	changed = 0
	i = length -1
	oldlist = contenders
	while i > 0:
		if contenders[i][0] not in contenders[i-1]:
			better = 2
			while better not in {0,1}:
				better = int(raw_input("Is %s better than %s? (1 for yes, 0 for no): "%(contenders[i][0],contenders[i-1][0])))
			if better == 0:
				contenders[i-1] += [contenders[i][0]]
			else:
				contenders[i] += [contenders[i-1][0]]
				temp = contenders [i-1]
				contenders [i-1] = contenders[i]
				contenders [i] = temp
				changed = 1
		i -= 1
	j += 1
	print j
	for i in contenders:
		cell = str(chr(64+j))+str((contenders.index(i)+1))
		ws[cell] = str(i[0])
for i in contenders:
	print i[0],
wb.save("bubble.xlsx")
